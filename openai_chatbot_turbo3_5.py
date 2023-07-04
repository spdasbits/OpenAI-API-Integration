import os
import openai
import gradio as gr
import socket
from openpyxl import load_workbook
import socket

openai.api_key = "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"

user_host_name = socket.gethostname()
start_sequence = "\nAI:"
restart_sequence = "\nHuman: "

question = "The following is a conversation with an AI assistant. The assistant is helpful, creative, clever, and very friendly.\n\nHuman: Hello, who are you?\nAI: I am an AI created by OpenAI. How can I help you today?\nHuman: "

def openai_create(question):
    workbook = load_workbook('context_to_chatgpt.xlsx')
    worksheet = workbook['Sales']
    cell_value = worksheet['A1'].value
    context=cell_value

    workbook = load_workbook('context_to_chatgpt.xlsx')
    worksheet = workbook['Sales']
    cell_value = worksheet['A1'].value
    context=cell_value
      
    worksheet = workbook['token_cost_cal']
    cell_value1 = worksheet['A2'].value
    last_total_token = (cell_value1)
      
    cell_value2 = worksheet['B2'].value
    last_total_token_cost = (cell_value2)
      
    context_token = (len(context)/3)
    context_token_cost = ((context_token/1000)*0.0015)

    response = openai.ChatCompletion.create(
    model="gpt-3.5-turbo",
    messages=[
        {"role": "assistant", "content": context},
        {"role": "user", "content": question}
    ]
    )
    output_token = (len(response.choices[0].message.content)/3)
    output_token_cost = (((output_token)/1000)*0.002)
    token = (context_token + output_token)
    cost = (context_token_cost + output_token_cost)
      
    last_total_token_final = (last_total_token + token)  
    last_total_token_cost_final = (last_total_token_cost + cost)

    workbook['token_cost_cal']['A2'] = last_total_token_final
    workbook['token_cost_cal']['B2'] = last_total_token_cost_final
    workbook.save('context_to_chatgpt.xlsx')
    return response.choices[0].message.content

def chatgpt_clone(input, history):
    history = history or []
    s = list(sum(history, ()))
    s.append(input)
    inp = ' '.join(s)
    output = openai_create(inp)
    history.append((input, output))
    return history, history


block = gr.Blocks()


with block:
    gr.Markdown(""+user_host_name+"""<h1><center>Hi Welcome to OpenAI ChatGPT (Note: Please do not use zf internal data)</center></h1>
    """)
    chatbot = gr.Chatbot()
    message = gr.Textbox(placeholder=question)
    state = gr.State()
    submit = gr.Button("SEND")
    submit.click(chatgpt_clone, inputs=[message, state], outputs=[chatbot, state])

block.launch(debug = True)
